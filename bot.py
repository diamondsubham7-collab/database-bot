import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from telegram import Update
from telegram.ext import Application, CommandHandler, MessageHandler, ContextTypes, filters
from dotenv import load_dotenv
import re
import zipfile
import io
import shutil
from datetime import datetime
import pdfplumber
from pypdf import PdfReader
import sqlite3
import threading
from http.server import HTTPServer, BaseHTTPRequestHandler
import tempfile
import time

load_dotenv()
TOKEN = os.getenv("BOT_TOKEN")
ADMIN_ID = int(os.getenv("ADMIN_ID", 0))
os.makedirs("received_files", exist_ok=True)

DB_PATH = "database.db"

# ---- GLOBAL FLAG FOR HEALTH CHECK ----
bot_is_alive = True

def init_db():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS users (
        user_id INTEGER PRIMARY KEY,
        username TEXT,
        first_name TEXT,
        last_name TEXT,
        join_date TEXT,
        last_active TEXT,
        total_actions INTEGER DEFAULT 0
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS logs (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER,
        command TEXT,
        details TEXT,
        timestamp TEXT
    )''')
    conn.commit()
    conn.close()

def log_action_sync(user_id, username, first_name, last_name, command, details=""):
    try:
        now = datetime.now().isoformat()
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute('''INSERT OR REPLACE INTO users 
                     (user_id, username, first_name, last_name, join_date, last_active, total_actions)
                     VALUES (?, ?, ?, ?, COALESCE((SELECT join_date FROM users WHERE user_id=?), ?), ?, 
                             COALESCE((SELECT total_actions FROM users WHERE user_id=?), 0) + 1)''',
                  (user_id, username, first_name, last_name, user_id, now, now, user_id))
        c.execute('''INSERT INTO logs (user_id, command, details, timestamp) VALUES (?, ?, ?, ?)''',
                  (user_id, command, details, now))
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"DB Error: {e}")

def beautify_excel(filepath, freeze_panes=False, auto_filter=False, alt_rows=False):
    try:
        wb = load_workbook(filepath)
        ws = wb.active
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_len:
                        max_len = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_len + 3
        if freeze_panes:
            ws.freeze_panes = 'A2'
        if auto_filter:
            ws.auto_filter.ref = ws.dimensions
        if alt_rows:
            fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                if row[0].row % 2 == 0:
                    for cell in row:
                        cell.fill = fill
        wb.save(filepath)
    except Exception as e:
        print(f"Beautify error: {e}")
    return filepath

def convert_to_excel(input_path, output_path=None):
    if output_path is None:
        output_path = os.path.splitext(input_path)[0] + ".xlsx"
    try:
        if input_path.endswith(".txt"):
            df = pd.read_csv(input_path, delimiter='\t', encoding='utf-8')
        else:
            df = pd.read_csv(input_path)
        if len(df.columns) == 1:
            df = pd.read_csv(input_path, sep=None, engine='python')
        df.to_excel(output_path, index=False, engine='openpyxl')
        beautify_excel(output_path, freeze_panes=True, auto_filter=True, alt_rows=True)
    except Exception as e:
        print(f"Conversion error: {e}")
        raise
    return output_path

# ---------- PDF TO EXCEL (FIXED) ----------
def pdf_to_excel(pdf_path, output_path=None):
    """Convert text-based PDF to Excel — robust parsing."""
    if output_path is None:
        output_path = os.path.splitext(pdf_path)[0] + ".xlsx"

    def smart_parse(lines):
        """Try every possible delimiter to parse lines into a DataFrame."""
        if not lines:
            return None

        # 1. Try comma (,)
        try:
            df = pd.read_csv(io.StringIO('\n'.join(lines)), engine='python', skipinitialspace=True)
            if len(df.columns) >= 2 and len(df) > 0:
                return df
        except:
            pass

        # 2. Try tab (\t)
        try:
            df = pd.read_csv(io.StringIO('\n'.join(lines)), delimiter='\t', engine='python', skipinitialspace=True)
            if len(df.columns) >= 2 and len(df) > 0:
                return df
        except:
            pass

        # 3. Try multiple spaces (2 or more)
        try:
            rows = []
            for line in lines:
                parts = re.split(r'\s{2,}', line.strip())
                if len(parts) > 1:
                    rows.append(parts)
            if rows:
                max_cols = max(len(r) for r in rows)
                for r in rows:
                    while len(r) < max_cols:
                        r.append('')
                if all(not re.search(r'\d', r[0]) for r in rows[:1]):
                    headers = rows[0]
                    data = rows[1:] if len(rows) > 1 else []
                else:
                    headers = [f"Col_{i+1}" for i in range(max_cols)]
                    data = rows
                df = pd.DataFrame(data, columns=headers)
                if len(df) > 0 and len(df.columns) > 1:
                    return df
        except:
            pass

        # 4. Key:Value format
        records = []
        cur = {}
        for line in lines:
            if ':' in line:
                k, v = line.split(':', 1)
                k, v = k.strip(), v.strip()
                if k == "Employee ID" and cur:
                    records.append(cur)
                    cur = {}
                cur[k] = v
        if cur:
            records.append(cur)
        if records:
            df = pd.DataFrame(records)
            if len(df) > 0 and len(df.columns) > 1:
                return df

        # 5. Single column fallback
        df = pd.DataFrame(lines, columns=["Data"])
        if len(df) > 0:
            return df
        return None

    try:
        all_text = []
        # Use pypdf first (most reliable for raw text)
        try:
            reader = PdfReader(pdf_path)
            for page in reader.pages:
                text = page.extract_text()
                if text:
                    all_text.append(text)
        except Exception as e:
            print(f"pypdf error: {e}")

        # Fallback to pdfplumber
        if not all_text:
            try:
                with pdfplumber.open(pdf_path) as pdf:
                    for page in pdf.pages:
                        text = page.extract_text()
                        if text:
                            all_text.append(text)
            except Exception as e:
                print(f"pdfplumber error: {e}")

        if not all_text:
            raise ValueError("❌ No selectable text found. Please upload a text-based PDF (not scanned).")

        full_text = "\n".join(all_text)
        lines = [line.strip() for line in full_text.split('\n') if line.strip()]
        df = smart_parse(lines)
        if df is None or len(df) == 0:
            raise ValueError("Could not parse PDF content into a table.")

        df.to_excel(output_path, index=False, engine='openpyxl')
        beautify_excel(output_path, freeze_panes=True, auto_filter=True, alt_rows=True)
        return output_path

    except Exception as e:
        print(f"PDF Error: {e}")
        raise

def parse_text_to_df(text):
    """Parse pasted CSV/TXT data."""
    lines = [line.strip() for line in text.split('\n') if line.strip()]
    if len(lines) < 2:
        return None
    delimiters = [',', '\t', ';', '|', r'\s+']
    for delim in delimiters:
        try:
            first_parts = re.split(delim, lines[0])
            if len(first_parts) < 2:
                continue
            counts = [len(re.split(delim, line)) for line in lines[:3]]
            if counts.count(counts[0]) >= 2:
                data_str = '\n'.join(lines)
                df = pd.read_csv(io.StringIO(data_str), delimiter=delim if delim != r'\s+' else None,
                                  engine='python', skipinitialspace=True)
                if len(df.columns) >= 2 and len(df) > 0:
                    return df
        except Exception:
            continue
    return None

# ---------- BOT COMMANDS ----------
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    log_action_sync(user.id, user.username or "", user.first_name or "", user.last_name or "", "/start", "User started bot")
    await update.message.reply_text(
        "📂 **Database Bot Ready!**\n\n"
        "📤 **Upload:**\n"
        "• TXT/CSV → Convert & Store\n"
        "• Excel → Store Directly\n"
        "• PDF → Extract & Convert to Excel (text-based only)\n"
        "• ZIP → Extract & Merge All (Auto)\n\n"
        "⚙️ **Commands:**\n"
        "/xlsx → Convert (No Store)\n"
        "/preview → Show Stored Files\n"
        "/merge → Merge All Files\n"
        "/append → Append New File\n"
        "/split [rows] → Split Large Excel\n"
        "/search [col] [value] → Search Records\n"
        "/filter [col] [op] [value] → Filter Data\n"
        "/stats → Show Statistics\n"
        "/report → Generate Summary Report (Excel)\n"
        "/clean → Remove Empty Rows/Cols\n"
        "/removeduplicate → Remove Duplicates\n"
        "/sort [col] → Sort by Column\n"
        "/clear → Delete All Files\n\n"
        "🤖 **AI Assistant:**\n"
        "• 'merge all files'\n"
        "• 'sort by salary'\n"
        "• 'search emp001'\n"
        "• 'filter department it'\n"
        "• 'remove duplicates'\n"
        "• 'generate report'\n\n"
        "💡 **Tip:** You can paste CSV/TXT data directly as a message!\n\n"
        "⚠️ **Important:** After your work is done, type **/clear** to delete stored files and free up server space.\n"
        "This ensures the bot remains fast and reliable for everyone."
    )

async def handle_document(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    file_path = None
    try:
        doc = update.message.document
        original_name = doc.file_name
        file_path = os.path.join("received_files", original_name)
        file = await doc.get_file()
        await file.download_to_drive(file_path)

        is_xlsx_mode = context.user_data.get('waiting_for_xlsx', False)
        is_append_mode = context.user_data.get('waiting_for_append', False)

        if original_name.endswith(".zip"):
            try:
                extract_path = os.path.join("received_files", f"zip_{update.effective_user.id}")
                os.makedirs(extract_path, exist_ok=True)
                with zipfile.ZipFile(file_path, 'r') as zip_ref:
                    zip_ref.extractall(extract_path)
                os.remove(file_path)

                all_files = []
                for root, dirs, files in os.walk(extract_path):
                    for f in files:
                        if f.endswith((".txt", ".csv")):
                            all_files.append(os.path.join(root, f))

                if not all_files:
                    shutil.rmtree(extract_path)
                    await update.message.reply_text("❌ No TXT or CSV files found in ZIP.")
                    return

                dfs = []
                for f in all_files:
                    if f.endswith(".txt"):
                        df = pd.read_csv(f, delimiter='\t', encoding='utf-8')
                    else:
                        df = pd.read_csv(f)
                    if len(df.columns) == 1:
                        df = pd.read_csv(f, sep=None, engine='python')
                    dfs.append(df)

                merged_df = pd.concat(dfs, ignore_index=True)
                merged_path = os.path.join("received_files", f"zip_merged_{update.effective_user.id}.xlsx")
                merged_df.to_excel(merged_path, index=False, engine='openpyxl')
                beautify_excel(merged_path, freeze_panes=True, auto_filter=True, alt_rows=True)

                shutil.rmtree(extract_path)

                await update.message.reply_document(document=open(merged_path, "rb"))
                os.remove(merged_path)
                await update.message.reply_text(f"✅ ZIP extracted! Merged {len(all_files)} files into 1 Excel.")
                log_action_sync(user.id, user.username or "", user.first_name or "", user.last_name or "", "ZIP Upload", f"Extracted and merged {len(all_files)} files")

            except Exception as e:
                await update.message.reply_text(f"⚠️ ZIP Error: {str(e)}")
                print(f"ZIP Error: {e}")
            finally:
                if os.path.exists(file_path):
                    os.remove(file_path)
            return

        if original_name.endswith(".pdf"):
            try:
                excel_path = pdf_to_excel(file_path)
                if is_xlsx_mode:
                    context.user_data['waiting_for_xlsx'] = False
                    await update.message.reply_document(document=open(excel_path, "rb"))
                    os.remove(excel_path)
                    await update.message.reply_text("✅ PDF → Excel (Not Stored).")
                    return
                if is_append_mode:
                    context.user_data['waiting_for_append'] = False
                    await append_file(update, context, excel_path, original_name)
                    return
                if 'files' not in context.user_data:
                    context.user_data['files'] = {}
                context.user_data['files'][original_name.replace('.pdf', '.xlsx')] = excel_path
                log_action_sync(user.id, user.username or "", user.first_name or "", user.last_name or "", "PDF Upload", f"Stored: {original_name}")
                await update.message.reply_text(f"✅ PDF converted and stored.")
            except Exception as e:
                await update.message.reply_text(f"⚠️ PDF Error: {str(e)}")
                print(f"PDF Error: {e}")
            finally:
                if os.path.exists(file_path):
                    os.remove(file_path)
            return

        if original_name.endswith((".txt", ".csv")):
            excel_path = convert_to_excel(file_path)
            if is_xlsx_mode:
                context.user_data['waiting_for_xlsx'] = False
                await update.message.reply_document(document=open(excel_path, "rb"))
                os.remove(excel_path)
                await update.message.reply_text("✅ Converted (Not Stored).")
                return
            if is_append_mode:
                context.user_data['waiting_for_append'] = False
                await append_file(update, context, excel_path, original_name)
                return
            if 'files' not in context.user_data:
                context.user_data['files'] = {}
            context.user_data['files'][original_name] = excel_path
            log_action_sync(user.id, user.username or "", user.first_name or "", user.last_name or "", "TXT/CSV Upload", f"Stored: {original_name}")
            await update.message.reply_text(f"✅ '{original_name}' Stored. ({len(context.user_data['files'])} files)")

        elif original_name.endswith((".xlsx", ".xls")):
            if is_xlsx_mode:
                context.user_data['waiting_for_xlsx'] = False
                await update.message.reply_document(document=open(file_path, "rb"))
                await update.message.reply_text("✅ File Sent (Not Stored).")
                return
            if is_append_mode:
                context.user_data['waiting_for_append'] = False
                await append_file(update, context, file_path, original_name)
                return
            if 'files' not in context.user_data:
                context.user_data['files'] = {}
            context.user_data['files'][original_name] = file_path
            log_action_sync(user.id, user.username or "", user.first_name or "", user.last_name or "", "Excel Upload", f"Stored: {original_name}")
            await update.message.reply_text(f"✅ '{original_name}' Stored. ({len(context.user_data['files'])} files)")

        elif original_name.lower().endswith(('.png', '.jpg', '.jpeg', '.bmp', '.tiff')):
            await update.message.reply_text("❌ Photo/OCR not supported on Render.")
            if os.path.exists(file_path):
                os.remove(file_path)

        else:
            await update.message.reply_text("❌ Only TXT, CSV, Excel, PDF, ZIP allowed.")
            if os.path.exists(file_path):
                os.remove(file_path)

    except Exception as e:
        await update.message.reply_text(f"⚠️ Error: {str(e)}")
        print(f"Error: {e}")
        if file_path and os.path.exists(file_path):
            try:
                os.remove(file_path)
            except:
                pass

async def append_file(update, context, new_file_path, original_name):
    stored = context.user_data.get('files', {})
    if not stored:
        if 'files' not in context.user_data:
            context.user_data['files'] = {}
        context.user_data['files'][original_name] = new_file_path
        await update.message.reply_text(f"✅ No stored files. '{original_name}' stored as new file.")
        return
    try:
        dfs = []
        for path in stored.values():
            dfs.append(pd.read_excel(path))
        dfs.append(pd.read_excel(new_file_path))
        merged_df = pd.concat(dfs, ignore_index=True)
        merged_path = os.path.join("received_files", f"merged_{update.effective_user.id}.xlsx")
        merged_df.to_excel(merged_path, index=False, engine='openpyxl')
        beautify_excel(merged_path, freeze_panes=True, auto_filter=True, alt_rows=True)
        for path in stored.values():
            try:
                os.remove(path)
            except:
                pass
        os.remove(new_file_path)
        context.user_data['files'] = {}
        context.user_data['files'][f"merged_{update.effective_user.id}.xlsx"] = merged_path
        await update.message.reply_document(document=open(merged_path, "rb"))
        await update.message.reply_text(f"✅ Appended '{original_name}' to existing data. Merged file sent and stored.")
    except Exception as e:
        await update.message.reply_text(f"⚠️ Append Error: {str(e)}")
        print(f"Append Error: {e}")

async def xlsx_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data['waiting_for_xlsx'] = True
    await update.message.reply_text("📤 Send TXT, CSV, PDF, or paste CSV/TXT data directly to convert (not stored).")

async def append_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    stored = context.user_data.get('files', {})
    if not stored:
        await update.message.reply_text("⚠️ No stored files. Upload some first.")
        return
    context.user_data['waiting_for_append'] = True
    await update.message.reply_text("📤 Send the file to append (TXT/CSV/PDF/Excel).")

async def preview_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    stored = context.user_data.get('files', {})
    if not stored:
        await update.message.reply_text("📭 No files stored.")
        return
    file_list = "\n".join([f"📄 {name}" for name in stored.keys()])
    await update.message.reply_text(f"📂 Stored Files ({len(stored)})\n{file_list}")

async def merge_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    stored = context.user_data.get('files', {})
    if len(stored) < 2:
        await update.message.reply_text("⚠️ Need at least 2 files to merge.")
        return
    try:
        dfs = []
        for path in stored.values():
            dfs.append(pd.read_excel(path))
        merged_df = pd.concat(dfs, ignore_index=True)
        merged_path = os.path.join("received_files", f"merged_{update.effective_user.id}.xlsx")
        merged_df.to_excel(merged_path, index=False, engine='openpyxl')
        beautify_excel(merged_path, freeze_panes=True, auto_filter=True, alt_rows=True)
        await update.message.reply_document(document=open(merged_path, "rb"))
        for path in stored.values():
            try:
                os.remove(path)
            except:
                pass
        context.user_data['files'] = {}
        os.remove(merged_path)
        await update.message.reply_text(f"✅ Merged {len(dfs)} files successfully!")
    except Exception as e:
        await update.message.reply_text(f"⚠️ Merge Error: {str(e)}")
        print(f"Merge Error: {e}")

async def split_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    stored = context.user_data.get('files', {})
    if not stored:
        await update.message.reply_text("📭 No files stored. Upload first.")
        return
    args = update.message.text.split()
    rows_per_file = 100
    if len(args) > 1:
        try:
            rows_per_file = int(args[1])
            if rows_per_file < 1:
                rows_per_file = 100
        except:
            rows_per_file = 100
    try:
        dfs = []
        for path in stored.values():
            dfs.append(pd.read_excel(path))
        merged_df = pd.concat(dfs, ignore_index=True)
        total_rows = len(merged_df)
        if total_rows == 0:
            await update.message.reply_text("⚠️ No data found.")
            return
        num_files = (total_rows + rows_per_file - 1) // rows_per_file
        if num_files == 1:
            await update.message.reply_text(f"⚠️ Only {total_rows} rows. Need more than {rows_per_file} to split.")
            return
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for i in range(num_files):
                start = i * rows_per_file
                end = min((i + 1) * rows_per_file, total_rows)
                chunk_df = merged_df.iloc[start:end]
                chunk_path = os.path.join("received_files", f"chunk_{i+1}_{update.effective_user.id}.xlsx")
                chunk_df.to_excel(chunk_path, index=False, engine='openpyxl')
                beautify_excel(chunk_path, freeze_panes=True, auto_filter=True, alt_rows=True)
                zip_file.write(chunk_path, f"part_{i+1}.xlsx")
                os.remove(chunk_path)
        zip_buffer.seek(0)
        await update.message.reply_document(
            document=zip_buffer,
            filename=f"split_files_{update.effective_user.id}.zip"
        )
        await update.message.reply_text(f"✅ Split {total_rows} rows into {num_files} files ({rows_per_file} rows each).")
    except Exception as e:
        await update.message.reply_text(f"⚠️ Split Error: {str(e)}")
        print(f"Split Error: {e}")

async def search_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    stored = context.user_data.get('files', {})
    if not stored:
        await update.message.reply_text("📭 No files stored.")
        return
    args = update.message.text.split(maxsplit=1)
    if len(args) < 2:
        await update.message.reply_text("❌ Usage: /search [column_name] [value]")
        return
    parts = args[1].rsplit(maxsplit=1)
    if len(parts) == 2:
        col_name = parts[0].strip()
        search_value = parts[1].strip()
    else:
        search_value = parts[0].strip()
        col_name = None
    try:
        dfs = []
        for path in stored.values():
            dfs.append(pd.read_excel(path))
        df = pd.concat(dfs, ignore_index=True)
        if col_name is None:
            col_name = df.columns[0]
        if col_name not in df.columns:
            await update.message.reply_text(f"❌ Column '{col_name}' not found.")
            return
        mask = df[col_name].astype(str).str.contains(search_value, case=False, na=False)
        result = df[mask]
        if len(result) == 0:
            await update.message.reply_text(f"❌ No records found.")
            return
        if len(result) > 10:
            msg = f"🔍 Found {len(result)} records (showing first 10):\n\n"
            result_display = result.head(10)
        else:
            msg = f"🔍 Found {len(result)} records:\n\n"
            result_display = result
        for idx, row in result_display.iterrows():
            msg += f"📌 **Row {idx+1}:**\n"
            for col in result_display.columns:
                msg += f"   • {col}: {row[col]}\n"
            msg += "\n"
        await update.message.reply_text(msg)
    except Exception as e:
        await update.message.reply_text(f"⚠️ Search Error: {str(e)}")

async def filter_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    stored = context.user_data.get('files', {})
    if not stored:
        await update.message.reply_text("📭 No files stored.")
        return
    args = update.message.text.split(maxsplit=1)
    if len(args) < 2:
        await update.message.reply_text("❌ Usage: /filter [column] [operator] [value]")
        return
    query = args[1].strip()
    operators = ['>=', '<=', '!=', '==', '>', '<', 'contains']
    operator = None
    col_name = None
    value = None
    for op in operators:
        if op in query:
            parts = query.split(op, 1)
            col_name = parts[0].strip()
            value = parts[1].strip()
            operator = op
            break
    if operator is None:
        parts = query.rsplit(maxsplit=1)
        if len(parts) == 2:
            col_name = parts[0].strip()
            value = parts[1].strip()
            operator = '=='
        else:
            await update.message.reply_text("❌ Invalid format.")
            return
    try:
        dfs = []
        for path in stored.values():
            dfs.append(pd.read_excel(path))
        df = pd.concat(dfs, ignore_index=True)
        if col_name not in df.columns:
            await update.message.reply_text(f"❌ Column '{col_name}' not found.")
            return
        if operator == 'contains':
            mask = df[col_name].astype(str).str.contains(value, case=False, na=False)
        elif operator == '>':
            mask = pd.to_numeric(df[col_name], errors='coerce') > float(value)
        elif operator == '<':
            mask = pd.to_numeric(df[col_name], errors='coerce') < float(value)
        elif operator == '>=':
            mask = pd.to_numeric(df[col_name], errors='coerce') >= float(value)
        elif operator == '<=':
            mask = pd.to_numeric(df[col_name], errors='coerce') <= float(value)
        elif operator == '!=':
            mask = df[col_name].astype(str) != value
        else:
            mask = df[col_name].astype(str) == value
        result = df[mask]
        if len(result) == 0:
            await update.message.reply_text(f"❌ No records found.")
            return
        if len(result) > 10:
            msg = f"📊 Found {len(result)} records (showing first 10):\n\n"
            result_display = result.head(10)
        else:
            msg = f"📊 Found {len(result)} records:\n\n"
            result_display = result
        for idx, row in result_display.iterrows():
            msg += f"📌 **Row {idx+1}:**\n"
            for col in result_display.columns:
                msg += f"   • {col}: {row[col]}\n"
            msg += "\n"
        await update.message.reply_text(msg)
    except Exception as e:
        await update.message.reply_text(f"⚠️ Filter Error: {str(e)}")

async def stats_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    stored = context.user_data.get('files', {})
    if not stored:
        await update.message.reply_text("📭 No files stored.")
        return
    try:
        dfs = []
        for path in stored.values():
            dfs.append(pd.read_excel(path))
        df = pd.concat(dfs, ignore_index=True)
        total = len(df)
        msg = f"📊 **Statistics**\n"
        msg += f"👥 Total Records: {total}\n"
        msg += f"📋 Total Columns: {len(df.columns)}\n"
        msg += f"📂 Columns: {', '.join(df.columns)}\n\n"
        msg += f"🔍 **Missing Values:**\n"
        for col in df.columns:
            missing = df[col].isna().sum()
            if missing > 0:
                msg += f"   • {col}: {missing} missing\n"
        if all(df[col].isna().sum() == 0 for col in df.columns):
            msg += "   ✅ No missing values found!\n"
        first_col = df.columns[0]
        dup_count = df[first_col].duplicated().sum()
        if dup_count > 0:
            msg += f"\n⚠️ Duplicate Records (based on '{first_col}'): {dup_count}\n"
        msg += f"\n📌 **Unique Values (Top 3 categorical):**\n"
        count = 0
        for col in df.columns:
            if df[col].dtype == 'object' and len(df[col].unique()) < 50:
                msg += f"   • {col}: {len(df[col].unique())} unique\n"
                count += 1
                if count >= 3:
                    break
        if count == 0:
            msg += "   (No categorical columns)\n"
        await update.message.reply_text(msg)
    except Exception as e:
        await update.message.reply_text(f"⚠️ Stats Error: {str(e)}")

async def report_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    stored = context.user_data.get('files', {})
    if not stored:
        await update.message.reply_text("📭 No files stored.")
        return
    try:
        dfs = []
        for path in stored.values():
            dfs.append(pd.read_excel(path))
        df = pd.concat(dfs, ignore_index=True)
        if len(df) == 0:
            await update.message.reply_text("⚠️ No data found.")
            return
        report_path = os.path.join("received_files", f"report_{update.effective_user.id}.xlsx")
        with pd.ExcelWriter(report_path, engine='openpyxl') as writer:
            overview_data = {
                'Metric': ['Total Records', 'Total Columns', 'Generated On'],
                'Value': [len(df), len(df.columns), datetime.now().strftime("%Y-%m-%d %H:%M:%S")]
            }
            overview_df = pd.DataFrame(overview_data)
            overview_df.to_excel(writer, sheet_name='Overview', index=False)
            preview_df = df.head(100)
            preview_df.to_excel(writer, sheet_name='Data Preview', index=False)
            missing_data = []
            for col in df.columns:
                missing_count = df[col].isna().sum()
                missing_percent = (missing_count / len(df)) * 100
                if missing_count > 0:
                    missing_data.append([col, missing_count, f"{missing_percent:.2f}%"])
            if missing_data:
                missing_df = pd.DataFrame(missing_data, columns=['Column', 'Missing Count', 'Missing %'])
            else:
                missing_df = pd.DataFrame([['No missing values', 0, '0%']], columns=['Column', 'Missing Count', 'Missing %'])
            missing_df.to_excel(writer, sheet_name='Missing Data', index=False)
            dept_col = None
            for col in df.columns:
                if col.lower() in ['department', 'dept', 'division']:
                    dept_col = col
                    break
            if dept_col:
                dept_counts = df[dept_col].value_counts().reset_index()
                dept_counts.columns = ['Department', 'Count']
                dept_counts.to_excel(writer, sheet_name='Department-wise', index=False)
            salary_col = None
            for col in df.columns:
                if 'salary' in col.lower() or 'wage' in col.lower():
                    try:
                        pd.to_numeric(df[col], errors='coerce')
                        salary_col = col
                        break
                    except:
                        pass
            if salary_col:
                salary_data = {
                    'Metric': ['Min Salary', 'Max Salary', 'Average Salary', 'Median Salary', 'Total Records'],
                    'Value': [
                        df[salary_col].min(),
                        df[salary_col].max(),
                        round(df[salary_col].mean(), 2),
                        df[salary_col].median(),
                        len(df)
                    ]
                }
                salary_df = pd.DataFrame(salary_data)
                salary_df.to_excel(writer, sheet_name='Salary Summary', index=False)
            first_col = df.columns[0]
            dup_count = df[first_col].duplicated().sum()
            dup_data = {
                'Column': [first_col],
                'Total Rows': [len(df)],
                'Duplicate Rows': [dup_count],
                'Unique Rows': [len(df) - dup_count]
            }
            dup_df = pd.DataFrame(dup_data)
            dup_df.to_excel(writer, sheet_name='Duplicates', index=False)
        beautify_excel(report_path, freeze_panes=True, auto_filter=True, alt_rows=True)
        await update.message.reply_document(document=open(report_path, "rb"))
        os.remove(report_path)
        await update.message.reply_text("✅ Report generated and sent!")
    except Exception as e:
        await update.message.reply_text(f"⚠️ Report Error: {str(e)}")
        print(f"Report Error: {e}")

async def clean_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    stored = context.user_data.get('files', {})
    if not stored:
        await update.message.reply_text("📭 No files stored.")
        return
    try:
        for name, path in stored.items():
            df = pd.read_excel(path)
            df = df.dropna(how='all')
            df = df.dropna(axis=1, how='all')
            df.to_excel(path, index=False, engine='openpyxl')
            beautify_excel(path, freeze_panes=True, auto_filter=True, alt_rows=True)
        await update.message.reply_text(f"🧹 Cleaned {len(stored)} files! Removed empty rows and columns.")
    except Exception as e:
        await update.message.reply_text(f"⚠️ Clean Error: {str(e)}")

async def removeduplicate_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    stored = context.user_data.get('files', {})
    if not stored:
        await update.message.reply_text("📭 No files stored.")
        return
    try:
        total_removed = 0
        for name, path in stored.items():
            df = pd.read_excel(path)
            first_col = df.columns[0]
            before = len(df)
            df_cleaned = df.drop_duplicates(subset=[first_col], keep='first')
            after = len(df_cleaned)
            removed = before - after
            total_removed += removed
            df_cleaned.to_excel(path, index=False, engine='openpyxl')
            beautify_excel(path, freeze_panes=True, auto_filter=True, alt_rows=True)
            if removed > 0:
                await update.message.reply_document(
                    document=open(path, "rb"),
                    filename=f"cleaned_{name}"
                )
        await update.message.reply_text(f"✅ Removed {total_removed} duplicate records from {len(stored)} files! Cleaned files sent.")
    except Exception as e:
        await update.message.reply_text(f"⚠️ Error: {str(e)}")

async def sort_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    stored = context.user_data.get('files', {})
    if not stored:
        await update.message.reply_text("📭 No files stored.")
        return
    args = update.message.text.split()
    col = args[1] if len(args) > 1 else None
    try:
        for name, path in stored.items():
            df = pd.read_excel(path)
            if col is None:
                col = df.columns[0]
            if col in df.columns:
                df = df.sort_values(by=col)
                df.to_excel(path, index=False, engine='openpyxl')
                beautify_excel(path, freeze_panes=True, auto_filter=True, alt_rows=True)
        await update.message.reply_text(f"✅ Sorted {len(stored)} files by '{col}'!")
    except Exception as e:
        await update.message.reply_text(f"⚠️ Sort Error: {str(e)}")

async def clear_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    stored = context.user_data.get('files', {})
    for path in stored.values():
        try:
            os.remove(path)
        except:
            pass
    context.user_data['files'] = {}
    await update.message.reply_text("🧹 All stored files cleared. Server space freed!")

async def admin_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    if user.id != ADMIN_ID:
        await update.message.reply_text("⛔ Unauthorized access.")
        return
    args = context.args
    if not args:
        await update.message.reply_text(
            "👑 **Admin Panel**\n\n"
            "/admin stats\n"
            "/admin users\n"
            "/admin logs [user_id]\n"
            "/admin broadcast [message]"
        )
        return
    subcommand = args[0].lower()
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    if subcommand == "stats":
        c.execute("SELECT COUNT(*) FROM users")
        total_users = c.fetchone()[0]
        c.execute("SELECT COUNT(*) FROM logs")
        total_actions = c.fetchone()[0]
        c.execute("SELECT COUNT(*) FROM logs WHERE date(timestamp) = date('now')")
        today_actions = c.fetchone()[0]
        await update.message.reply_text(
            f"📊 **Bot Statistics**\n"
            f"👥 Total Users: {total_users}\n"
            f"📝 Total Actions: {total_actions}\n"
            f"📈 Today's Actions: {today_actions}"
        )
    elif subcommand == "users":
        c.execute("SELECT user_id, username, first_name, last_name, total_actions, last_active FROM users ORDER BY total_actions DESC LIMIT 20")
        rows = c.fetchall()
        if not rows:
            await update.message.reply_text("No users found.")
        else:
            msg = "👥 **Top 20 Users**\n\n"
            for row in rows:
                name = row[2] or row[1] or str(row[0])
                msg += f"• {name} (ID: {row[0]}) - Actions: {row[4]} - Last: {row[5][:10]}\n"
            await update.message.reply_text(msg)
    elif subcommand == "logs":
        target_user = int(args[1]) if len(args) > 1 else None
        if target_user:
            c.execute("SELECT command, details, timestamp FROM logs WHERE user_id = ? ORDER BY id DESC LIMIT 20", (target_user,))
        else:
            c.execute("SELECT user_id, command, details, timestamp FROM logs ORDER BY id DESC LIMIT 20")
        rows = c.fetchall()
        if not rows:
            await update.message.reply_text("No logs found.")
        else:
            msg = "📜 **Recent Logs**\n\n"
            for row in rows:
                if target_user:
                    msg += f"• {row[0]} | {row[1]} | {row[2]} | {row[3][:16]}\n"
                else:
                    msg += f"• User {row[0]} | {row[1]} | {row[2]} | {row[3][:16]}\n"
            await update.message.reply_text(msg)
    elif subcommand == "broadcast":
        message = " ".join(args[1:])
        if not message:
            await update.message.reply_text("❌ Please provide a message.")
            return
        c.execute("SELECT user_id FROM users")
        users = c.fetchall()
        sent = 0
        for u in users:
            try:
                await context.bot.send_message(chat_id=u[0], text=f"📢 **Announcement**\n\n{message}")
                sent += 1
            except:
                pass
        await update.message.reply_text(f"✅ Broadcast sent to {sent} users.")
    conn.close()

async def handle_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    text = update.message.text.strip()
    is_xlsx_mode = context.user_data.get('waiting_for_xlsx', False)

    df = parse_text_to_df(text)
    if df is not None:
        if is_xlsx_mode:
            context.user_data['waiting_for_xlsx'] = False
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
                tmp_path = tmp.name
            df.to_excel(tmp_path, index=False, engine='openpyxl')
            beautify_excel(tmp_path, freeze_panes=True, auto_filter=True, alt_rows=True)
            await update.message.reply_document(document=open(tmp_path, "rb"))
            os.remove(tmp_path)
            await update.message.reply_text("✅ Pasted data converted to Excel (Not Stored).")
            return
        else:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"pasted_data_{timestamp}.xlsx"
            file_path = os.path.join("received_files", filename)
            df.to_excel(file_path, index=False, engine='openpyxl')
            beautify_excel(file_path, freeze_panes=True, auto_filter=True, alt_rows=True)
            if 'files' not in context.user_data:
                context.user_data['files'] = {}
            context.user_data['files'][filename] = file_path
            log_action_sync(user.id, user.username or "", user.first_name or "", user.last_name or "", "Paste Data", f"Stored as {filename}")
            await update.message.reply_text(f"✅ Pasted data stored as '{filename}'. ({len(context.user_data['files'])} files)")
            return

    text_clean = text.lower().strip('.,?!')
    try:
        if "merge" in text_clean or "combine" in text_clean:
            await merge_command(update, context)
        elif "clear" in text_clean or "delete all" in text_clean or "remove all" in text_clean:
            await clear_command(update, context)
        elif "stat" in text_clean or "overview" in text_clean and "report" not in text_clean:
            await stats_command(update, context)
        elif "report" in text_clean or "generate report" in text_clean:
            await report_command(update, context)
        elif "clean" in text_clean or "remove empty" in text_clean:
            await clean_command(update, context)
        elif "remove duplicate" in text_clean or "deduplicate" in text_clean:
            await removeduplicate_command(update, context)
        elif "split" in text_clean:
            nums = re.findall(r'\d+', text_clean)
            if nums:
                update.message.text = f"/split {nums[0]}"
            else:
                update.message.text = "/split"
            await split_command(update, context)
        elif "sort by" in text_clean or "sort" in text_clean:
            parts = text_clean.split("by")
            if len(parts) > 1:
                col = parts[1].strip().split()[0] if parts[1].strip() else None
                if col:
                    update.message.text = f"/sort {col}"
                else:
                    update.message.text = "/sort"
            else:
                update.message.text = "/sort"
            await sort_command(update, context)
        elif "search" in text_clean or "find" in text_clean or "show" in text_clean:
            parts = text_clean.split()
            keywords = ["search", "find", "show", "for", "me", "details", "record", "of"]
            for word in keywords:
                if word in parts:
                    parts.remove(word)
            if parts:
                value = " ".join(parts).strip()
                update.message.text = f"/search {value}"
                await search_command(update, context)
            else:
                await update.message.reply_text("❌ Please specify what to search.")
        elif "filter" in text_clean:
            query = text_clean.replace("filter", "").strip()
            if query:
                update.message.text = f"/filter {query}"
                await filter_command(update, context)
            else:
                await update.message.reply_text("❌ Please specify filter.")
        else:
            if is_xlsx_mode:
                await update.message.reply_text("❌ I couldn't parse your message as CSV/TXT data. Please send a file or paste valid data.\n\nExample:\nEmployeeID,Name,Salary\nEMP001,Rahul,45000")
            else:
                await update.message.reply_text(
                    "🤖 **I didn't understand.**\n\n"
                    "Try: 'merge all files', 'sort by salary', 'search emp001', 'filter department it', 'remove duplicates', 'generate report', 'clean data', 'split 100', 'stats', 'clear all'\n\n"
                    "💡 You can also paste CSV/TXT data directly!"
                )
    except Exception as e:
        await update.message.reply_text(f"⚠️ AI Error: {str(e)}")
        print(f"AI Error: {e}")

# ---------- MAIN WITH SMART HEALTH CHECK ----------
def main():
    global bot_is_alive
    bot_is_alive = True

    # ---- Health Check Server ----
    class HealthHandler(BaseHTTPRequestHandler):
        def do_GET(self):
            global bot_is_alive
            if bot_is_alive:
                self.send_response(200)
                self.send_header('Content-type', 'text/plain')
                self.end_headers()
                self.wfile.write(b'OK')
            else:
                self.send_response(503)
                self.send_header('Content-type', 'text/plain')
                self.end_headers()
                self.wfile.write(b'Service Unavailable: Bot polling stopped')

        def log_message(self, format, *args):
            return  # Disable spam logs

    def run_health_server():
        try:
            port = int(os.environ.get('PORT', 8080))
            server = HTTPServer(('0.0.0.0', port), HealthHandler)
            server.serve_forever()
        except Exception as e:
            print(f"❌ Health server crashed: {e}")
            os._exit(1)

    health_thread = threading.Thread(target=run_health_server, daemon=False)
    health_thread.start()
    print(f"✅ Health check server running on port {os.environ.get('PORT', 8080)}")

    # ---- Bot Polling Thread ----
    def run_bot():
        global bot_is_alive
        try:
            init_db()
            app = Application.builder().token(TOKEN).build()
            app.add_handler(CommandHandler("start", start))
            app.add_handler(CommandHandler("xlsx", xlsx_command))
            app.add_handler(CommandHandler("preview", preview_command))
            app.add_handler(CommandHandler("merge", merge_command))
            app.add_handler(CommandHandler("append", append_command))
            app.add_handler(CommandHandler("split", split_command))
            app.add_handler(CommandHandler("search", search_command))
            app.add_handler(CommandHandler("filter", filter_command))
            app.add_handler(CommandHandler("stats", stats_command))
            app.add_handler(CommandHandler("report", report_command))
            app.add_handler(CommandHandler("clean", clean_command))
            app.add_handler(CommandHandler("removeduplicate", removeduplicate_command))
            app.add_handler(CommandHandler("sort", sort_command))
            app.add_handler(CommandHandler("clear", clear_command))
            app.add_handler(CommandHandler("admin", admin_command))
            app.add_handler(MessageHandler(filters.Document.ALL, handle_document))
            app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text))

            print("✅ Bot polling started. Waiting for messages...")
            app.run_polling()
            print("⚠️ Bot polling stopped unexpectedly.")
            bot_is_alive = False
            os._exit(1)

        except Exception as e:
            print(f"❌ Bot polling CRASHED: {e}")
            bot_is_alive = False
            os._exit(1)

    bot_thread = threading.Thread(target=run_bot, daemon=False)
    bot_thread.start()

    # ---- Keep main thread alive ----
    try:
        while True:
            time.sleep(10)
            if not bot_thread.is_alive():
                print("⚠️ Bot thread died! Setting health check to fail...")
                bot_is_alive = False
                os._exit(1)
    except KeyboardInterrupt:
        print("Shutting down...")
        os._exit(0)

if __name__ == "__main__":
    main()